from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from openpyxl import Workbook, load_workbook

from repositories.banking_repository import BankingRepository
from repositories.category_repository import CategoryRepository


class ExcelService:
    """Importiert und exportiert Buchungen als Excel-Datei."""
    REQUIRED_HEADERS = ["Wertstellung", "Betrag", "Empfänger", "Verwendungszweck"]

    def __init__(self, banking_repository=None, category_repository=None):
        self.banking = banking_repository or BankingRepository()
        self.categories = category_repository or CategoryRepository()

    def export_account(self, customer_id, account_id):
        if not self.banking.account_belongs_to_user(account_id, customer_id):
            raise PermissionError("Auf dieses Finanzkonto besteht kein Zugriff.")
        account = self.banking.get_account(account_id, customer_id)
        entries = self.banking.list_all_entries_for_export(account_id)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Buchungen"
        sheet.append(self.REQUIRED_HEADERS + ["Kategorie"])

        for entry in entries:
            sheet.append(
                [
                    entry["wertstellungsdatum"],
                    entry["betrag"],
                    entry["empfaenger"],
                    entry["verwendungszweck"],
                    entry["kategorie"],
                ]
            )
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 28
        sheet.column_dimensions["D"].width = 45
        sheet.column_dimensions["E"].width = 24

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"{account['kontoname'].replace(' ', '_')}_export.xlsx"
        return output, filename

    def import_account(self, customer_id, account_id, file_storage):
        if not self.banking.account_belongs_to_user(account_id, customer_id):
            raise PermissionError("Auf dieses Finanzkonto besteht kein Zugriff.")
        if not file_storage or not file_storage.filename:
            raise ValueError("Bitte eine Excel-Datei auswählen.")
        if not file_storage.filename.lower().endswith(".xlsx"):
            raise ValueError("Es werden nur Dateien im Format .xlsx unterstützt.")

        workbook = load_workbook(file_storage, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if headers[:4] != self.REQUIRED_HEADERS:
            raise ValueError(
                "Die ersten vier Spalten müssen Wertstellung, Betrag, Empfänger und Verwendungszweck heißen."
            )

        imported = 0
        skipped = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            value_date, amount, recipient, purpose = row[:4]
            if all(value is None for value in (value_date, amount, recipient, purpose)):
                continue
            try:
                if isinstance(value_date, datetime):
                    parsed_date = value_date
                else:
                    parsed_date = datetime.fromisoformat(str(value_date))
                parsed_amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                if parsed_amount == 0:
                    raise ValueError
                recipient = str(recipient or "").strip()
                purpose = str(purpose or "").strip()
                if not recipient or not purpose:
                    raise ValueError
            except (ValueError, InvalidOperation, TypeError):
                skipped.append(row_number)
                continue

            match = self.categories.find_matching_category(customer_id, recipient, purpose)
            category_id = (
                match["kategorieid"]
                if match
                else self.categories.get_uncategorized_id(customer_id)
            )
            self.banking.create_entry(
                account_id=account_id,
                category_id=category_id,
                value_date=parsed_date,
                amount=parsed_amount,
                recipient=recipient,
                purpose=purpose,
            )
            imported += 1

        return imported, skipped
